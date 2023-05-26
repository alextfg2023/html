import os
import mysql.connector
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Ruta al directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Ruta al archivo de Excel
excel_file = os.path.join(script_dir, 'calendarios/calendarios_escolares/2023/calendario.xlsx')
output_file = os.path.join(script_dir, 'resultado.txt')

# Establecer la conexión con la base de datos de MySQL
conn = mysql.connector.connect(
    host='localhost',
    user='Timer_Lab',
    password='TimerLab2023.',
    database='timerlab'
)
cursor = conn.cursor()

try:
    # Borrar los datos existentes en la tabla
    delete_query = "DELETE FROM calendario_escolar"
    cursor.execute(delete_query)

    # Cargar el libro de Excel
    wb = load_workbook(excel_file)

    # Seleccionar la hoja de cálculo
    sheet = wb.active

    # Iterar sobre las filas de la hoja de cálculo y ejecutar las consultas de inserción
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fecha = row[0].strftime('%Y-%m-%d')
        descripcion = row[3]
        tipo = row[2]
        dia_semana = row[1]
        query = f"INSERT INTO calendario_escolar (fecha, descripcion, tipo, dia_semana) VALUES ('{fecha}', '{descripcion}', '{tipo}', '{dia_semana}')"
        cursor.execute(query)

    # Confirmar los cambios y cerrar la conexión
    conn.commit()
    print("Los datos se han insertado correctamente en la tabla.")

    # Crear y escribir en el archivo de salida
    with open(output_file, 'w') as file:
        file.write("La ejecución del script 'Calendario_escolar' ha finalizado correctamente.")

except Exception as e:
    # En caso de error, realizar un rollback y mostrar el mensaje de error
    conn.rollback()
    print("Ha ocurrido un error durante la inserción de datos:")
    print(str(e))
    # Crear y escribir en el archivo de salida
    with open(output_file, 'w') as file:
        file.write(f"Ha ocurrido un error durante la inserción de datos:\n{str(e)}")

finally:
    cursor.close()
    conn.close()
